'''
worbook = planilha
sheet = página
'''

import openpyxl

# Criação de workbook
workbook = openpyxl.Workbook()

# Mostrar sheets existentes
print(workbook.sheetnames)

# Criar nova sheet
workbook.create_sheet('ruas')
workbook.create_sheet('cidades')
workbook.create_sheet('estados')

# Salvar modificações para uma planilha
workbook.save('endereços.xlsx')


# Alterar o nome de um sheet
workbook['ruas'].title = 'ruas da cidade'
workbook.save('endereços.xlsx')

# Como exlcuir um sheet
del workbook['Sheet']
print(workbook.sheetnames)
workbook.save('endereços.xlsx')